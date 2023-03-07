import re
import time

from django.shortcuts import render, HttpResponse, redirect
from django import forms
from app01 import models
from api import views
from django.db.models import Q
from django.forms.models import model_to_dict
from django.conf import settings

# Create your views here.

DF='default_avatar.png'
a = {}


def layout(request):
    return render(request, "layout.html")
def login(request):
    """
    #Administrtor.objects.create(admin_id="root",password="hebei12#$",app="yes")
    Administrtor.objects.filter(admin_id="root").update\
        (
            naminate_privilege="yes",
            add_user_privilege ="yes",
            manage_user_privilege ="yes",
            manage_message_privilege ="yes",
            audit_privilege ="yes",
        )
    """
    if request.method == "GET":

        return render(request, "login.html")
    else:
        # print(request.POST)
        print(request.POST.get("user"), request.POST.get("pwd"))
        user = request.POST.get("user")
        pwd = request.POST.get("pwd")
        check = models.Administrtor.objects.filter(admin_id=user, password=pwd).first()
        print(check)
        if check == None:
            return render(request, "login.html", {"error_msg": "账号或密码错误"})
        elif user == check.admin_id and pwd == check.password:
            request.session["info"] = {
                "id":check.id,
                "admin_id":check.admin_id,
                "naminate_privilege":check.naminate_privilege,
                "add_user_privilege":check.add_user_privilege,
                "audit_privilege":check.audit_privilege,
                "manage_message_privilege":check.manage_message_privilege,
                "manage_user_privilege":check.manage_user_privilege

            }

            return redirect("/index/")

        else:
            return render(request, "login.html", {"error_msg": "账号或密码错误"})
        # return HttpResponse("登录成功")
    # else:
def logout(request):
    request.session.clear()
    return redirect('/login/')
def index(request):
    if request.session.get('info') == None:
        return redirect('/login/')
    ad_id = request.session.get('info')
    adminuser = models.Administrtor.objects.filter(admin_id=ad_id.get('admin_id')).first()
    return render(request, "index.html", {"adminuser": adminuser})


def admin_list(request):

    #return HttpResponse({"s":"sdfa"})

    print(request.session.get('info'))
    if request.session.get('info') == None:
        return redirect('/login/')
    ad_id = request.session.get('info')
    adminuser = models.Administrtor.objects.filter(admin_id=ad_id.get('admin_id')).first()
    print(adminuser.naminate_privilege)
    print(adminuser.add_user_privilege)
    print(adminuser.audit_privilege)
    print(adminuser.manage_message_privilege)
    print(adminuser.manage_user_privilege)
    fil  = Q(naminate_privilege__lte=adminuser.naminate_privilege)
    if adminuser.add_user_privilege == False:
        if fil == '':
            fil = Q(add_user_privilege=False)
        else:
            fil = fil&Q(add_user_privilege=False)
    if adminuser.audit_privilege == False:
        if fil == '':
            fil = Q(audit_privilege=False)
        else:
            fil = fil&Q(audit_privilege=False)
    if adminuser.manage_message_privilege == False:
        if fil == '':
            fil = Q(manage_message_privilege=False)
        else:
            fil = fil&Q(manage_message_privilege=False)
    if adminuser.manage_user_privilege == False:
        if fil == '':
            fil = Q(manage_user_privilege=False)
        else:
            fil = fil&Q(manage_user_privilege=False)

    ass = Q(naminate_privilege__lte=adminuser.naminate_privilege)
    bb = Q(audit_privilege=False)
    cc = ass&bb
    print(cc)

    print(Q(naminate_privilege__lte=adminuser.naminate_privilege)&Q(audit_privilege=False))
    print(11, fil)
    ass = models.Administrtor.objects.filter(fil).all().order_by("-naminate_privilege")
    print(ass)

    return render(request,"admin_list.html",{"ass":ass,"adminuser":adminuser})
def admin_add(request):
    if request.session.get('info') == None:
        return redirect('/login/')

    ad_id = request.session.get('info')
    adminuser = models.Administrtor.objects.filter(admin_id=ad_id.get('admin_id')).first()
    numlevel = range(0,adminuser.naminate_privilege)
    #print('get', numlevel)
    if request.method == "GET":

        return render(request, "admin_add.html",{"lis":adminuser,"num":numlevel})

    #print(request.POST)
    user = request.POST.get("user")
    pwd = request.POST.get("pwd")
    grade = request.POST.get("grade")
    add_user= bool(request.POST.get("add_user"))
    manage_user= bool(request.POST.get("manage_user"))
    manage_message= bool(request.POST.get("manage_message"))
    audit= bool(request.POST.get("audit"))
    # print(user)
    # print(pwd)
    # print(grade)
    # print(user!='' and pwd != '' and add_user|manage_user|manage_message|audit )
    #
    # print(user!='')
    # print(pwd != '')
    if user!='' and pwd != '' and add_user|manage_user|manage_message|audit :
        dc = models.Administrtor.objects.filter(admin_id=user).first()
        if dc == None:


            print("dc",dc)
            models.Administrtor.objects.create\
                (
                    admin_id=user,
                    password=pwd,
                    naminate_privilege=grade,
                    add_user_privilege=add_user,
                    manage_user_privilege=manage_user,
                    manage_message_privilege=manage_message,
                    audit_privilege=audit,
                )
            return redirect("/admin/list/")
        else:
            dc = "用户名被占用"
    else:
        dc = "用户名或密码不能为空，权限至少一个"


    return render(request, "admin_add.html",{"lis":adminuser,"num":numlevel,"dc":dc})
def admin_editpwd(request,nid):
    if request.session.get('info') == None:
        return redirect('/login/')
    if request.method == "GET":
         return render(request, "admin_editpwd.html")
    ad_id = request.session.get('info')
    print(request.POST)
    print(request.POST.get('pwd_old'))
    print(request.POST.get('pwdnew1'))
    print(request.POST.get('pwdnew2'))
    if request.POST.get('pwd_old') and request.POST.get('pwdnew1') and request.POST.get('pwdnew2'):
        print("fkong")
        dc = models.Administrtor.objects.filter(id = nid ,admin_id=ad_id.get('admin_id'),password=request.POST.get('pwd_old')).first()
        if dc == None:
            error="管理员密码不正确"
        else:
            if request.POST.get('pwdnew1') == request.POST.get('pwdnew2'):
                print("上传数据库")
                models.Administrtor.objects.filter(id = nid ,admin_id=ad_id.get('admin_id'),password=request.POST.get('pwd_old')).update(password=request.POST.get('pwdnew2'))
                error=""
                return redirect("/index/")
            else:
                error = "两次密码不一致"
    else:
        print("kong")
        error = "所有空不能为空"


    #adminuser = models.Administrtor.objects.filter(admin_id=ad_id.get('admin_id')).first()
    return render(request,"admin_editpwd.html",{"error":error})
def admin_edit(request,nid):
    if request.session.get('info') == None:
        return redirect('/login/')
    ad_id = request.session.get('info')
    adminuser = models.Administrtor.objects.filter(admin_id=ad_id.get('admin_id')).first()
    numlevel = range(0,adminuser.naminate_privilege)
    print(nid)
    if request.method == "GET":
        user_text = models.Administrtor.objects.filter(id=nid).first()

        print(user_text.add_user_privilege)
        print(user_text.audit_privilege)
        print(user_text.manage_message_privilege)
        print(user_text.manage_user_privilege)
        checked = {"checked1":None,"checked2":None,"checked3":None,"checked4":None,}
        if user_text.add_user_privilege == True:
            print("checket")
            checked["checked1"] = 'checked'
        if user_text.manage_user_privilege == True:
            print("checket")
            checked["checked2"] = 'checked'
        if user_text.manage_message_privilege == True:
            print("checket")
            checked["checked3"] = 'checked'
        if user_text.audit_privilege == True:
            print("checket")
            checked["checked4"] = 'checked'


        return render(request, 'admin_edit.html',{"lis":adminuser,"num":numlevel,"checked":checked,"user_text":user_text})
    user = request.POST.get("user")
    pwd = request.POST.get("pwd")
    grade = request.POST.get("grade")
    add_user = bool(request.POST.get("add_user"))
    manage_user = bool(request.POST.get("manage_user"))
    manage_message = bool(request.POST.get("manage_message"))
    audit = bool(request.POST.get("audit"))
    print(user)
    print(pwd)
    print(grade)
    print(user!='' and pwd != '' and add_user|manage_user|manage_message|audit )
    print(user!='')
    print(pwd != '')

    if user != '' and pwd != '' and add_user | manage_user | manage_message | audit:

        dc = models.Administrtor.objects.filter(id=nid).first()
        if dc == None:
            dc = "查无此人"
        else:
            print("dc", dc)
            models.Administrtor.objects.filter(id=nid).update \
                    (
                    admin_id=user,
                    password=pwd,
                    naminate_privilege=grade,
                    add_user_privilege=add_user,
                    manage_user_privilege=manage_user,
                    manage_message_privilege=manage_message,
                    audit_privilege=audit,
                )
            return redirect("/admin/list/")

    else:
        dc = "用户名或密码不能为空，权限至少一个"

    return render(request, "admin_add.html", {"lis": adminuser, "num": numlevel, "dc": dc})
def admin_delete(request, nid):
    if request.session.get('info') == None:
        return redirect('/login/')
    print(nid)
    models.Administrtor.objects.filter(id=nid).delete()
    # #return HttpResponse("asdafdsafdas")
    return redirect("/admin/list/")
    #return HttpResponse("sfasd")

def user_list(request):
    if request.session.get('info') == None:
        return redirect('/login/')
    #models.CommentRecord.objects.filter().delete()
    #models.Admin_send.objects.create(title=True,title_type='message')
    # models.NewsMessage.objects.filter().delete()
    # models.Task.objects.filter().delete()
    # models.UserInfo.objects.filter(student_number="210809039040").update(avatarurl="user/210809039040.png")
    data_list = models.UserInfo.objects.all()
    # print(data_list)
    return render(request, "user_list.html", {"data_list": data_list})
def user_add(request):
    global user_info
    if request.session.get('info') == None:
        return redirect('/login/')
    # global error
    if request.method == "POST":
        user_info = {
            "id": request.POST.get("student_number"),
            "name": request.POST.get("name"),
            "gender": request.POST.get("gender"),
            "age": request.POST.get("age"),
            "student_number": request.POST.get("student_number"),
            "password": request.POST.get("student_number"),
            "create_time": request.POST.get("time") + "-01",
            "level": models.UserLevel.objects.get(id=request.POST.get("level")),
            "campus_id": request.POST.get("campus"),
            "department_id": request.POST.get("department"),
            "profession_id": request.POST.get("profession"),
            "username": request.POST.get("student_number"),
            "avatarurl": DF,

        }

    campus = models.Campus.objects.all()
    department = models.Department.objects.all()
    profession = models.Profession.objects.all()
    level = models.UserLevel.objects.all()

    if request.method == "GET":
        return render(request, 'user_add.html',
                      {"campus": campus, "department": department, "profession": profession, "level": level})
    if not request.POST.get("name"):
        error = "用户名不能为空"
        return render(request, 'user_add.html',
                      {"campus": campus, "department": department, "profession": profession, "level": level,
                       "error": error})
    a = request.POST.get("student_number")
    print('a', a)
    res = re.findall('[1-9]\d*', a)
    print('res', res[0])
    print()
    if a == res[0]:
        if request.POST.get("time"):
            if not models.UserInfo.objects.filter(username=request.POST.get("username")):
                if not models.UserInfo.objects.filter(student_number=request.POST.get("student_number")):
                    models.UserInfo.objects.create(**user_info)
                    print(user_info)
                    return redirect("/user/list/")
                else:
                    error = "学号重复"
            else:
                error = "用户名重复"
        else:
            error = "日期不能为空"
    else:
        error = "学号必须为纯数字"
    return render(request, 'user_add.html',
                  {"campus": campus, "department": department, "profession": profession, "level": level,
                   "error": error})
def user_upload(request):
    from openpyxl import load_workbook
    print(request.POST)
    print(request.FILES.get('file'))
    file_object = request.FILES.get('file')
    wb = load_workbook(file_object)
    print(wb)
    sheet = wb.worksheets[0]
    # print(sheet.cell(1,2).value)
    # cell=sheet.cell(1,1)
    userlist=[]
    for row in sheet.iter_rows(min_row=2):
        print(len(row))
        userdict = {}
        for x in range(0,len(row)):
            #print(sheet.iter_rows[0].row[x].value)
            #print(sheet.cell(1,x+1).value+':'+row[x].value)

            userdict[sheet.cell(1,x+1).value] = row[x].value
            if sheet.cell(1,x+1).value == 'student_number':
                userdict['id'] = row[x].value
                userdict['username'] = row[x].value
                userdict['password'] = row[x].value

            if sheet.cell(1,x+1).value == 'level':
                userdict.update({sheet.cell(1, x + 1).value:models.UserLevel.objects.get(id=row[x].value)})
        userdict['avatarurl'] = DF

        #print(time.strftime("%Y-%m-%d",time.gmtime()))
        userdict['create_time'] = time.strftime("%Y-%m-%d",time.gmtime())
        print(userdict['id'])
        one = models.UserInfo.objects.filter(id = userdict['id'])
        print(one.first())
        if one.first() == None:
        #print(row[1].value)
            models.UserInfo.objects.create(**userdict)


    return redirect('/user/list/')
def user_edit(request, nid):
    if request.session.get('info') == None:
        return redirect('/login/')
    campus = models.Campus.objects.all()
    department = models.Department.objects.all()
    profession = models.Profession.objects.all()
    level = models.UserLevel.objects.all()
    gender_01 = ""
    gender_02 = ""


    user_text = models.UserInfo.objects.filter(id=nid).first()
    user = model_to_dict(user_text)
    #print(user['avatarurl'])
    image_name = str(user['avatarurl'])
    avatar = views.URL + 'media/user/' +str(user['avatarurl'])
    # print(avatar)
    user['avatarurl'] = {'avatar':avatar,'image':image_name}
    #print(user)


    if user_text.gender == 1:
        gender_01 = "checked"
    else:
        gender_02 = "checked"
        #print(user_text.create_time)
    creat_time = re.findall('[0-9]\d*', str( user_text.create_time))
    if request.method == "GET":
        return render(request, 'user_edit.html', {"campus":campus,"department": department,"profession": profession,"level": level,"user_text": user,"gender_01": gender_01,"gender_02": gender_02,"creat_time":creat_time})
    print(request.POST)
    if request.POST.get('avatarurl') == '1':
        avatarurl = DF
    else:
        avatarurl = request.POST.get('avatarurl')
    #print(avatarurl)
    # if request.POST.get('username') == '':
    #     print('kong')
    i=0
    for x in request.POST:
        #print(request.POST.get(x))
        if request.POST.get(x) == '':
            i = i+1
    #print(i)
    if i !=0:
        return render(request, 'user_edit.html',
                      {"error":"所有空不能为空","campus": campus, "department": department, "profession": profession, "level": level,
                       "user_text": user, "gender_01": gender_01, "gender_02": gender_02, "creat_time": creat_time})

    # file_object = request.FILES.get('avatarurl')
    # img_name = request.POST.get("username")+".png"
    # b = '%s/user/%s' % (settings.MEDIA_ROOT, img_name)
    # f = open( b,mode='wb')
    # for chunk in file_object.chunks():#chunk相当于文件块一点一点上传
    #     f.write(chunk)      #读取一点写一点
    # f.close()
    # img_url = 'http://192.168.42.111:15050'+settings.MEDIA_URL+'user/'+img_name
    # print(img_url)
    #
    user_info = {
        "id":request.POST.get("student_number"),
        "name": request.POST.get("name"),
        "gender": request.POST.get("gender"),
        "age": request.POST.get("age"),
        "student_number": request.POST.get("student_number"),
        "password": request.POST.get("student_number"),
        "create_time": request.POST.get("time") + "-01",
        "level": models.UserLevel.objects.get(id=request.POST.get("level")),
        "campus_id": request.POST.get("campus"),
        "department_id": request.POST.get("department"),
        "profession_id": request.POST.get("profession"),
        "username": request.POST.get("username"),
        "avatarurl": avatarurl,

    }
    models.UserInfo.objects.filter(id=nid).update(**user_info)
    #

    return redirect("/user/list/")
def user_delete(request, nid):
    if request.session.get('info') == None:
        return redirect('/login/')
    models.UserInfo.objects.filter(id=nid).delete()
    # #return HttpResponse("asdafdsafdas")
    return redirect("/user/list/")


def audit_list(request):
    if request.session.get('info') == None:
        return redirect('/login/')

    ad_id = request.session.get('info')
    admin_audit = models.Administrtor.objects.filter(admin_id=ad_id.get('admin_id')).first()
    #print(adminuser.audit_privilege)
    if not admin_audit.audit_privilege:
        print('无权限')
        return redirect('/index/')
    adminsend = models.Admin_send.objects.filter(Q(title_type="baidu_text")|Q(title_type="baidu_image")|Q(title_type="baidu_video")).all()
    #print(adminsend)
    false_list = models.NewsMessage.objects.filter(~Q(admin_hide="True")).all()
    print(false_list)
    return render(request,"audit_list.html",{"adminsend":adminsend,"list":false_list,"admin_audit":admin_audit})
def audit_detail(request,nid):
    if request.session.get('info') == None:
        return redirect('/login/')

    ad_id = request.session.get('info')
    admin_audit = models.Administrtor.objects.filter(admin_id=ad_id.get('admin_id')).first()
    #print(adminuser.audit_privilege)
    if not admin_audit.audit_privilege:
        print('无权限')
        return redirect('/index/')
    print(nid)
    one = models.NewsMessage.objects.filter(id=nid)
    print(one.first().type)
    type_text =''
    if one.first() != None:

        # from api import views
        # ser = views.WX_indexModelSerializer(instance=one,many=True)
        #print(ser.data)

        if one.get().type == 'text':
            # print("text")
            dict_text = models.Task.objects.filter(time_TS=one.get().time_TS, stu_id=one.get().type_text)
            print(dict_text.get().text)
            #return dict_text.get().text
            type_text = dict_text.get().text

        if one.get().type == 'image':
            # print('image')
            dict_image = models.Task.objects.filter(time_TS=one.get().time_TS, stu_id=one.get().type_text)
            # print(len(dict_image))
            num = 0
            assds = []
            for asses in dict_image:

                #if asses.hide != 'True':
                    # print(asses.hide)
                assds.append({'url': asses.text,'id': num})


                num = num + 1

            print(assds)
            type_text = assds

        if one.get().type == "video":
            dict_video = models.Task.objects.filter(time_TS=one.get().time_TS, stu_id=one.get().type_text).all()
            #print(dict_video)
            video_dif = {}
            for ins in dict_video:
                # print(ins)
                #if ins.hide != 'True':
                    # assds.append({'url': ins.text})
                    # print(ins.text[-4:])
                if ins.text[-4:] == '.mp4':
                        # print('ssssssssssssssssssssss')
                    video_dif["video_url"] = ins.text
                else:
                    video_dif["video_image"] = ins.text
            print(video_dif)
            type_text = video_dif


    return render(request,"audit_detail.html",{
        "obj":one.first(),
        "ser":type_text
    })
    #return HttpResponse('sdfadf')
def audit_baidu(request,nid):
    #print("13413413413",nid)
    baidu = models.Admin_send.objects.filter(id=nid).first()
    #print(baidu.title)
    if baidu.title == "True":
        models.Admin_send.objects.filter(id=nid).update(title = "False")
    else:
        models.Admin_send.objects.filter(id=nid).update(title = "True")


    return redirect('/audit/list/')
    #return HttpResponse("sfa")
def audit_pass(request,nid):
    print(nid)


    mess_type = models.NewsMessage.objects.filter(id=nid).get()
    #print(mess_type)
    admin_hide = models.NewsMessage.objects.filter(id=nid).get().admin_hide
    #print(admin_hide)

    if mess_type.type == 'text':

        print('text格式')
        if admin_hide != 'True':
            models.NewsMessage.objects.filter(time_TS=mess_type.time_TS, type_text=mess_type.type_text).update(admin_hide ='True')
            models.Task.objects.filter(time_TS=mess_type.time_TS, stu_id=mess_type.type_text).update(hide ='True')
            return redirect('/audit/list/')
        else:
            models.NewsMessage.objects.filter(time_TS=mess_type.time_TS, type_text=mess_type.type_text).update(admin_hide ='False')
            models.Task.objects.filter(time_TS=mess_type.time_TS, stu_id=mess_type.type_text).update(hide ='False')
            return redirect('/manage/list/')

    elif mess_type.type == 'image':
        imglength = 0
        print('image格式')
        if admin_hide != 'True':
            image_list = models.Task.objects.filter(time_TS=mess_type.time_TS, stu_id=mess_type.type_text).all()
            # print('自动审核', len(image_list))
            if len(image_list)==0:
                # print('列表为空')
                return
            for one in image_list:
                if one.hide != 'True':
                    models.Task.objects.filter(id=one.id,time_TS=mess_type.time_TS, stu_id=mess_type.type_text).update(hide='True')
                    imglength = imglength +1
            if imglength == len(image_list):
                models.NewsMessage.objects.filter(time_TS=mess_type.time_TS, type_text=mess_type.type_text).update(admin_hide ='True')
            return redirect('/audit/list/')
        else:
            image_list = models.Task.objects.filter(time_TS=mess_type.time_TS, stu_id=mess_type.type_text).all()
            # print('自动审核', len(image_list))
            if len(image_list)==0:
                # print('列表为空')
                return
            for one in image_list:
                if one.hide != 'False':
                    models.Task.objects.filter(id=one.id,time_TS=mess_type.time_TS, stu_id=mess_type.type_text).update(hide='False')
                    imglength = imglength +1
            if imglength == len(image_list):
                models.NewsMessage.objects.filter(time_TS=mess_type.time_TS, type_text=mess_type.type_text).update(admin_hide ='False')
            return redirect('/manage/list/')

    elif mess_type.type == 'video':
        print('video格式')
        global videolen
        videolen= 0
        if admin_hide != 'True':
            video_list = models.Task.objects.filter(time_TS=mess_type.time_TS, stu_id=mess_type.type_text).all()
            # print('自动审核', len(video_list))
            if len(video_list) < 2:
                # print('列表为空')
                return
            for video_one in video_list:
                if video_one.hide != 'True':
                    models.Task.objects.filter(id=video_one.id, time_TS=video_one.time_TS, stu_id=video_one.stu_id).update(hide='True')
            models.NewsMessage.objects.filter(time_TS=mess_type.time_TS, type_text=mess_type.type_text).update(admin_hide ='True')
            return redirect('/audit/list/')
        else:
            video_list = models.Task.objects.filter(time_TS=mess_type.time_TS, stu_id=mess_type.type_text).all()
            # print('自动审核', len(video_list))
            if len(video_list) < 2:
                # print('列表为空')
                return
            for video_one in video_list:
                if video_one.hide != 'False':
                    models.Task.objects.filter(id=video_one.id, time_TS=video_one.time_TS, stu_id=video_one.stu_id).update(hide='False')
            models.NewsMessage.objects.filter(time_TS=mess_type.time_TS, type_text=mess_type.type_text).update(admin_hide ='False')
            return redirect('/manage/list/')
def audit_del(request,nid):
    print(nid)
    ass = models.NewsMessage.objects.filter(id=nid).first()
    print(ass.admin_hide)
    if ass.admin_hide == 'True':
        back = '/manage/list/'
    else:
        back = '/audit/list/'
    models.NewsMessage.objects.filter(id=nid).delete()
    return redirect(back)
def manage_list(request):
    if request.session.get('info') == None:
        return redirect('/login/')

    ad_id = request.session.get('info')
    admin_audit = models.Administrtor.objects.filter(admin_id=ad_id.get('admin_id')).first()
    #print(adminuser.audit_privilege)
    if not admin_audit.manage_message_privilege:
        print('无权限')
        return redirect('/index/')
    #adminsend = models.Admin_send.objects.filter(Q(title_type="baidu_text")|Q(title_type="baidu_image")|Q(title_type="baidu_video")).all()
    #print(adminsend)
    false_list = models.NewsMessage.objects.filter(Q(admin_hide="True")).all()
    print(false_list)





    return render(request,"manage_list.html",{
        #"adminsend":adminsend,
        "list":false_list,"admin_audit":admin_audit})




#     #Administrtor.objects.all().delete()
#
#     if request.method == "GET":
#         return render(request, "admin_add.html")
#
#     user = request.POST.get("user")
#     pwd = request.POST.get("pwd")
#     grade = request.POST.get("grade")
#     add_user= request.POST.get("add_user")
#     manage_user= request.POST.get("manage_user")
#     manage_message= request.POST.get("manage_message")
#     audit= request.POST.get("audit")
#     print(user)
#     print(pwd)
#     print(grade)
#     print(add_user)
#     print(manage_user)
#     print(manage_message)
#     print(audit)
#     models.Administrtor.objects.create\
#         (
#             admin_id=user,
#             password=pwd,
#             naminate_privilege=grade,
#             add_user_privilege=add_user,
#             manage_user_privilege=manage_user,
#             manage_message_privilege=manage_message,
#             audit_privilege=audit,
#         )
#     return redirect("/index/")
# def delete (request):
#     nid = request.GET.get("nid")
#     models.Administrtor.objects.filter(admin_id=nid).delete()
#     return redirect("/index/")
#
#
#
# class UserModelForm(forms.ModelForm):
#     name = forms.CharField(min_length=3,label="姓名")
#     class Meta:
#         model = models.UserInfo
#         fields = ["name","gender","age","student_number","password","campus","department","profession","create_time","level"]
#         # widgets = {
#         #     "name":forms.TextInput(attrs={"class":"layui-input"})
#         # }
#     def __init__(self, *args, **kwargs):
#         super().__init__(*args, **kwargs)
#         for name, field in self.fields.items():
#             print(name,field)
#             field.widget.attrs = {"class": "layui-input"}
#
#
#
# def user_edit(request,nid):
#     row_object = models.UserInfo.objects.filter(id=nid).first()
#     if request.method =="GET":
#         form = UserModelForm(instance=row_object)
#         return render(request,"user_edit.html",{"form":form})
#     form = UserModelForm(data=request.POST,instance=row_object)
#     if form.is_valid():
#         # 默认保存的是用户输入的所有数据，如果想要再用户输入以外增加一点值
#         # form. instance .字段名=值
#         print(form.cleaned_data)
#
#         form.save()
#         return redirect("/index/")
#     else:
#
#         print(form.cleaned_data)
#         print(form.errors)
#         return render(request,"user_edit.html",{"from":form})

#
